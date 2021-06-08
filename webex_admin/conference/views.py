import datetime
import json
import mimetypes
import os

import openpyxl as openpyxl
import requests

from django.http import HttpResponse
from django.shortcuts import render, redirect

from Dashboard.models import Details
from openpyxl import Workbook
from openpyxl.styles import Font

from .models import *
from webex_admin.settings import CLIENT_ID, CLIENT_SECRET

from webex_admin.settings import BASE_DIR


def webex_panel(request):
    if request.user.is_authenticated:
        try:
            details_object = Details.objects.filter(teacher__teacher=request.user).get()
        except:
            return redirect('/dashboard')

        teams_list = ClassRoom.objects.all()
        teams = []
        for team in teams_list:
            teams.append({
                'title': team.title,
                'meeting_id': team.meeting_id,
                'meeting_number': team.meeting_number,
            })

        data = {
            'details': {
                'teacher': str(details_object.teacher.id),
                'id': str(details_object.id),
                'name': details_object.name,
                'profile_pic': {
                    'status': 1 if bool(details_object.profile_file) else 0,
                    'url': details_object.profile_file.url if bool(details_object.profile_file) else '',
                }
            },
            'teams': teams
        }
        return render(request, "conference/webex_panel.html", {'data': data})
    return redirect('/')


def team_edit(request, meeting_id):
    if request.method == 'GET':
        if request.user.is_authenticated:
            try:
                details_object = Details.objects.filter(teacher__teacher=request.user).get()
                token_data = AccessToken.objects.filter(user=request.user).get()
                if datetime.datetime.now(datetime.timezone.utc) - token_data.last_modification >= datetime.timedelta(days=10):
                    headers = {'Content-Type': 'application/json'}
                    url = "https://api.ciscospark.com/v1/access_token"
                    body = {
                        'grant_type': 'refresh_token',
                        'client_id': CLIENT_ID,
                        'client_secret': CLIENT_SECRET,
                        'refresh_token': token_data.refresh_token
                    }
                    response = requests.post(url, headers=headers, data=json.dumps(body)).json()
                    token_data.token = response['token']
                    token_data.refresh_token = response['refresh_token']
                    token_data.save()

                headers = {
                    'Authorization': f'Bearer {token_data.token}',
                    'Content-Type': 'application/json'
                }
                team = ClassRoom.objects.filter(meeting_id=meeting_id).get()

                url = f"https://webexapis.com/v1/team/memberships?teamId={team.cisco_id}"
                response = requests.get(url, headers=headers).json()
                response = response['items']
                members = []
                for member in response:
                    members.append({
                        'membership_id': member['id'],
                        'display_name': member['personDisplayName'],
                        'email': member['personEmail'],
                    })
            except:
                return redirect('/dashboard')

            data = {
                'details': {
                    'teacher': str(details_object.teacher.id),
                    'id': str(details_object.id),
                    'name': details_object.name,
                    'profile_pic': {
                        'status': 1 if bool(details_object.profile_file) else 0,
                        'url': details_object.profile_file.url if bool(details_object.profile_file) else '',
                    }
                },
                'meeting_id': meeting_id,
                'members': members
            }
            return render(request, "conference/team_edit.html", {'data': data})
        return redirect('/')


def team_create(request):
    if request.method == 'GET':
        if request.user.is_authenticated:
            try:
                details_object = Details.objects.filter(teacher__teacher=request.user).get()
            except:
                return redirect('/dashboard')

            data = {
                'details': {
                    'teacher': str(details_object.teacher.id),
                    'id': str(details_object.id),
                    'name': details_object.name,
                    'profile_pic': {
                        'status': 1 if bool(details_object.profile_file) else 0,
                        'url': details_object.profile_file.url if bool(details_object.profile_file) else '',
                    }
                },
            }
            return render(request, "conference/team_create.html", {'data': data})
        return redirect('/')
    elif request.method == 'POST':
        if request.user.is_authenticated:
            try:
                details_object = Details.objects.filter(teacher__teacher=request.user).get()
            except:
                return redirect('/dashboard')

            title = request.POST.get('title')
            if title is None or title == '':
                message = "Invaild Title"
            else:
                try:
                    token_data = AccessToken.objects.filter(user=request.user).get()
                    if datetime.datetime.now(datetime.timezone.utc) - token_data.last_modification >= datetime.timedelta(
                            days=10):
                        headers = {'Content-Type': 'application/json'}
                        url = "https://api.ciscospark.com/v1/access_token"
                        body = {
                            'grant_type': 'refresh_token',
                            'client_id': CLIENT_ID,
                            'client_secret': CLIENT_SECRET,
                            'refresh_token': token_data.refresh_token
                        }
                        response = requests.post(url, headers=headers, data=json.dumps(body)).json()
                        token_data.token = response['token']
                        token_data.refresh_token = response['refresh_token']
                        token_data.save()

                    headers = {
                        'Authorization': f'Bearer {token_data.token}',
                        'Content-Type': 'application/json'
                    }

                    url = "https://webexapis.com/v1/teams"
                    body = {
                        'name': title
                    }
                    response = requests.post(url, headers=headers, data=json.dumps(body)).json()

                    team = ClassRoom.objects.create(id=uuid.uuid4(), title=response['name'], cisco_id=response['id'])

                    url = f"https://webexapis.com/v1/rooms?teamId={team.cisco_id}"
                    response = requests.get(url, headers=headers).json()
                    team.room_id = response['items'][0]['id']
                    team.save()

                    url = f"https://webexapis.com/v1/rooms/{team.room_id}/meetingInfo"
                    response = requests.get(url, headers=headers).json()
                    team.meeting_number = response['meetingNumber']
                    team.save()

                    url = f"https://webexapis.com/v1/meetings?meetingNumber={response['meetingNumber']}"
                    response = requests.get(url, headers=headers).json()
                    team.meeting_id = response['items'][0]['id']
                    team.save()
                    message = "Team Created Successfully"
                except:
                    try:
                        team.delete()
                    except:
                        message = "Something Problem to Create Team, Try again..!"
                    message = "Something Problem to Create Team, Try again..!"

            data = {
                'details': {
                    'teacher': str(details_object.teacher.id),
                    'id': str(details_object.id),
                    'name': details_object.name,
                    'profile_pic': {
                        'status': 1 if bool(details_object.profile_file) else 0,
                        'url': details_object.profile_file.url if bool(details_object.profile_file) else '',
                    }
                },
                'message': message
            }
            return render(request, "conference/team_create.html", {'data': data})
        return redirect('/')
    else:
        return HttpResponse('<h1>400 Not Found Error</h1>')


def team_attendance(request, meeting_id):
    if request.user.is_authenticated:
        if ClassRoom.objects.filter(meeting_id=meeting_id).count() == 0:
            return HttpResponse('<h1>400 Not Found Error</h1>')
        else:
            try:
                team = ClassRoom.objects.filter(meeting_id=meeting_id).get()

                token_data = AccessToken.objects.filter(user=request.user).get()
                if datetime.datetime.now(datetime.timezone.utc) - token_data.last_modification >= datetime.timedelta(
                        days=10):
                    headers = {'Content-Type': 'application/json'}
                    url = "https://api.ciscospark.com/v1/access_token"
                    body = {
                        'grant_type': 'refresh_token',
                        'client_id': CLIENT_ID,
                        'client_secret': CLIENT_SECRET,
                        'refresh_token': token_data.refresh_token
                    }
                    response = requests.post(url, headers=headers, data=json.dumps(body)).json()
                    token_data.token = response['token']
                    token_data.refresh_token = response['refresh_token']
                    token_data.save()

                headers = {
                    'Authorization': f'Bearer {token_data.token}',
                    'Content-Type': 'application/json'
                }

                url = f"https://webexapis.com/v1/teams/{team.cisco_id}"
                response = requests.get(url, headers=headers).json()
                title = response['name']

                url = f"https://webexapis.com/v1/meetingParticipants?meetingId={meeting_id}"
                response = requests.get(url, headers=headers).json()
                response = response['items']
            except:
                return HttpResponse('<h1>No Participant Found</h1>')

            wb = openpyxl.Workbook()
            sheet = wb.active

            cell = sheet.cell(row=1, column=1)
            cell.value = "Id"
            cell.font = Font(bold=True)

            cell = sheet.cell(row=1, column=2)
            cell.value = "Name"
            cell.font = Font(bold=True)

            cell = sheet.cell(row=1, column=3)
            cell.value = "Email"
            cell.font = Font(bold=True)

            cell = sheet.cell(row=1, column=4)
            cell.value = "State"
            cell.font = Font(bold=True)

            cell = sheet.cell(row=1, column=5)
            cell.value = "Joined Time"
            cell.font = Font(bold=True)

            sheet.column_dimensions['A'].width = 10
            sheet.column_dimensions['B'].width = 40
            sheet.column_dimensions['C'].width = 50
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 40

            for row, attendee in enumerate(response, start=2):
                cell = sheet.cell(row=row, column=1)
                cell.value = str(int(row)-1)

                cell = sheet.cell(row=row, column=2)
                cell.value = str(attendee['displayName'])

                cell = sheet.cell(row=row, column=3)
                cell.value = str(attendee['email'])

                cell = sheet.cell(row=row, column=4)
                cell.value = str(attendee['state'])

                cell = sheet.cell(row=row, column=5)
                epoch_time = int(attendee['devices'][0]['joinedTime'])
                cell.value = str(datetime.datetime.fromtimestamp(epoch_time/1000))

            now = datetime.datetime.now()
            filename = str(title) + '_D_' + str(now.strftime("%Y")) + '_' + str(now.strftime("%m")) + '_' + str(now.strftime("%d")) + '_T_' + str(now.strftime("%H:%M:%S")).replace(':', '-') + "_Meet_" + str(meeting_id) + '.xlsx'
            filepath = os.path.join(BASE_DIR, str(r'../../attendance/' + filename))
            wb.save(filepath)

            file = open(filepath, 'rb')
            mime_type, _ = mimetypes.guess_type(filepath)
            response = HttpResponse(file, content_type=mime_type)
            response['Content-Disposition'] = "attachment; filename=%s" % filename
            return response

    return redirect('/')


def team_delete(request, meeting_id):
    if request.user.is_authenticated:
        try:
            token_data = AccessToken.objects.filter(user=request.user).get()
            if datetime.datetime.now(datetime.timezone.utc) - token_data.last_modification >= datetime.timedelta(days=10):
                headers = {'Content-Type': 'application/json'}
                url = "https://api.ciscospark.com/v1/access_token"
                body = {
                    'grant_type': 'refresh_token',
                    'client_id': CLIENT_ID,
                    'client_secret': CLIENT_SECRET,
                    'refresh_token': token_data.refresh_token
                }
                response = requests.post(url, headers=headers, data=json.dumps(body)).json()
                token_data.token = response['token']
                token_data.refresh_token = response['refresh_token']
                token_data.save()

            headers = {
                'Authorization': f'Bearer {token_data.token}',
                'Content-Type': 'application/json'
            }
            team = ClassRoom.objects.filter(meeting_id=meeting_id).get()

            url = f"https://webexapis.com/v1/teams/{team.cisco_id}"
            requests.delete(url, headers=headers).json()

            team.delete()
            return redirect('/webex')
        except:
            return HttpResponse('<h1>No Team Found</h1>')


def team_member_delete(request, membership_id):
    if request.user.is_authenticated:
        try:
            token_data = AccessToken.objects.filter(user=request.user).get()
            if datetime.datetime.now(datetime.timezone.utc) - token_data.last_modification >= datetime.timedelta(days=10):
                headers = {'Content-Type': 'application/json'}
                url = "https://api.ciscospark.com/v1/access_token"
                body = {
                    'grant_type': 'refresh_token',
                    'client_id': CLIENT_ID,
                    'client_secret': CLIENT_SECRET,
                    'refresh_token': token_data.refresh_token
                }
                response = requests.post(url, headers=headers, data=json.dumps(body)).json()
                token_data.token = response['token']
                token_data.refresh_token = response['refresh_token']
                token_data.save()

            headers = {
                'Authorization': f'Bearer {token_data.token}',
                'Content-Type': 'application/json'
            }
            url = f"https://webexapis.com/v1/team/memberships/{membership_id}"
            requests.delete(url, headers=headers)

            return redirect('/webex')
        except:
            return HttpResponse('<h1>No Member Found</h1>')


def team_member_add(request, meeting_id):
    if request.method == 'POST':
        if request.user.is_authenticated:
            name = request.POST.get('name')
            email = request.POST.get('email')
            class_year = request.POST.get('class_year')
            section = request.POST.get('section')
            roll_no = request.POST.get('roll_no')
            moderator = True if request.POST.get('moderator') == 'true' else False

            try:
                token_data = AccessToken.objects.filter(user=request.user).get()
                if datetime.datetime.now(datetime.timezone.utc) - token_data.last_modification >= datetime.timedelta(
                        days=10):
                    headers = {'Content-Type': 'application/json'}
                    url = "https://api.ciscospark.com/v1/access_token"
                    body = {
                        'grant_type': 'refresh_token',
                        'client_id': CLIENT_ID,
                        'client_secret': CLIENT_SECRET,
                        'refresh_token': token_data.refresh_token
                    }
                    response = requests.post(url, headers=headers, data=json.dumps(body)).json()
                    token_data.token = response['token']
                    token_data.refresh_token = response['refresh_token']
                    token_data.save()

                headers = {
                    'Authorization': f'Bearer {token_data.token}',
                    'Content-Type': 'application/json'
                }
                url = f"https://webexapis.com/v1/people?email={email}"
                response = requests.get(url, headers=headers).json()
                response = response['items']

                if StudentData.objects.filter(name=name, email=email, class_year=class_year, section=section, roll_no=roll_no).count() == 0:
                    display_name = str(class_year) + '_' + str(section) + '_' + str(roll_no) + '_' + str(name)
                    StudentData.objects.create(id=uuid.uuid4(), name=name, email=email, class_year=class_year, section=section, roll_no=roll_no, cisco_id=response[0]['id'], display_name=display_name)

                team = ClassRoom.objects.filter(meeting_id=meeting_id).get()
                url = f"https://webexapis.com/v1/team/memberships"
                body = {
                    'teamId': team.cisco_id,
                    'personId': response[0]['id'],
                    'isModerator': moderator
                }
                requests.post(url, headers=headers, data=json.dumps(body)).json()

                return redirect('/webex/team/' + meeting_id + '/edit')
            except:
                return HttpResponse('<h1>No Account Found</h1>')
        return redirect('/')
    return HttpResponse('<h1>404 Not Found</h1>')


def team_member_add_list(request, meeting_id):
    if request.method == 'POST':
        if request.user.is_authenticated:

            file = request.FILES.get('file')
            try:
                token_data = AccessToken.objects.filter(user=request.user).get()
                if datetime.datetime.now(datetime.timezone.utc) - token_data.last_modification >= datetime.timedelta(
                        days=10):
                    headers = {'Content-Type': 'application/json'}
                    url = "https://api.ciscospark.com/v1/access_token"
                    body = {
                        'grant_type': 'refresh_token',
                        'client_id': CLIENT_ID,
                        'client_secret': CLIENT_SECRET,
                        'refresh_token': token_data.refresh_token
                    }
                    response = requests.post(url, headers=headers, data=json.dumps(body)).json()
                    token_data.token = response['token']
                    token_data.refresh_token = response['refresh_token']
                    token_data.save()

                headers = {
                    'Authorization': f'Bearer {token_data.token}',
                    'Content-Type': 'application/json'
                }
            except:
                return render(request, "error/index.html",
                              {'error': 'Something Went Wrong',
                               'message': "Invalid Token"})

            if file.content_type not in ['.csv', 'application/vnd.ms-excel',
                                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
                return render(request, "error/index.html",
                              {'error': 'Invalid File',
                               'message': "Wrong File Format"})

            workbook = openpyxl.load_workbook(file)
            worksheet = workbook.active

            file_data = list()
            for row in worksheet.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                file_data.append(row_data)

            for student in file_data:
                name = student[0]
                email = student[1]
                class_year = student[2]
                section = student[3]
                roll_no = student[4]
                moderator = True if student[5] == '1' else False
                try:
                    url = f"https://webexapis.com/v1/people?email={email}"
                    response = requests.get(url, headers=headers).json()
                    response = response['items']

                    if StudentData.objects.filter(name=name, email=email, class_year=class_year, section=section, roll_no=roll_no).count() == 0:
                        display_name = str(class_year) + '_' + str(section) + '_' + str(roll_no) + '_' + str(name)
                        StudentData.objects.create(id=uuid.uuid4(), name=name, email=email, class_year=class_year, section=section, roll_no=roll_no, cisco_id=response[0]['id'], display_name=display_name)

                    team = ClassRoom.objects.filter(meeting_id=meeting_id).get()
                    url = f"https://webexapis.com/v1/team/memberships"
                    body = {
                        'teamId': team.cisco_id,
                        'personId': response[0]['id'],
                        'isModerator': moderator
                    }
                    requests.post(url, headers=headers, data=json.dumps(body)).json()

                    return redirect('/webex/team/' + meeting_id + '/edit')
                except:
                    return HttpResponse('<h1>No Account Found</h1>')
        return redirect('/')
    return HttpResponse('<h1>404 Not Found</h1>')


def team_admit(request, meeting_id):
    if request.user.is_authenticated:
        try:
            team = ClassRoom.objects.filter(meeting_id=meeting_id).get()

            token_data = AccessToken.objects.filter(user=request.user).get()
            if datetime.datetime.now(datetime.timezone.utc) - token_data.last_modification >= datetime.timedelta(
                    days=10):
                headers = {'Content-Type': 'application/json'}
                url = "https://api.ciscospark.com/v1/access_token"
                body = {
                    'grant_type': 'refresh_token',
                    'client_id': CLIENT_ID,
                    'client_secret': CLIENT_SECRET,
                    'refresh_token': token_data.refresh_token
                }
                response = requests.post(url, headers=headers, data=json.dumps(body)).json()
                token_data.token = response['token']
                token_data.refresh_token = response['refresh_token']
                token_data.save()

            headers = {
                'Authorization': f'Bearer {token_data.token}',
                'Content-Type': 'application/json'
            }

            url = f"https://webexapis.com/v1/team/memberships?teamId={team.cisco_id}"
            response = requests.get(url, headers=headers).json()
            response = response['items']
            members = []
            for member in response:
                members.append(member['personEmail'])

            url = f"https://webexapis.com/v1/meetingParticipants?meetingId={meeting_id}"
            response = requests.get(url, headers=headers).json()
            participants_list = response['items']

            for participant in participants_list:
                if participant['email'] in members:
                    if participant['state'] == 'lobby':
                        url = f"https://webexapis.com/v1/meetingParticipants/{participant['id']}"
                        body = {
                            'admit': True
                        }
                        requests.put(url, headers=headers, data=json.dumps(body)).json()

            return redirect('/webex')
        except:
            return render(request, "error/index.html",
                          {'error': 'Something Went Wrong',
                           'message': "Invalid Meeting ID"})
    else:
        return redirect('/')
